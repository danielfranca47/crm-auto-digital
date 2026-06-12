"""Export de leads para ficheiro Excel (.xlsx)."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


_COLUMNS = [
    ("name",          "Nome"),
    ("phone",         "Telefone"),
    ("website",       "Website"),
    ("address",       "Endereço"),
    ("rating",        "Avaliação"),
    ("reviews_count", "Nº Avaliações"),
    ("maps_url",      "Link Google Maps"),
]

_HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def export_to_excel(
    items: List[Dict[str, Any]],
    query: str,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Exporta lista de leads para .xlsx.
    Se output_path for None, salva em ~/Downloads/leads_<timestamp>.xlsx.
    Retorna o Path do ficheiro criado.
    """
    if output_path is None:
        import datetime
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        downloads = Path.home() / "Downloads"
        downloads.mkdir(exist_ok=True)
        output_path = downloads / f"leads_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Cabeçalho de info
    ws["A1"] = f"Pesquisa: {query}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Total: {len(items)} leads"
    ws["A2"].font = Font(size=10, color="6B7280")
    ws.append([])  # linha vazia

    # Cabeçalhos de coluna (linha 4)
    header_row = 4
    for col_idx, (_, label) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for item in items:
        row = []
        for key, _ in _COLUMNS:
            val = item.get(key)
            row.append(val if val is not None else "")
        ws.append(row)

    # Larguras automáticas
    _col_widths = {
        "name": 35, "phone": 18, "website": 35, "address": 45,
        "rating": 12, "reviews_count": 14, "maps_url": 50,
    }
    for col_idx, (key, _) in enumerate(_COLUMNS, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _col_widths.get(key, 20)

    # Altura do cabeçalho
    ws.row_dimensions[header_row].height = 22

    # Freeze panes na linha de dados
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(output_path)
    return output_path
